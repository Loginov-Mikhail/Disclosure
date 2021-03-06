# *****************************************************************************
#   Программа disclosure.py предназначена для поиска на сайте
# www.e-disclosure.ru новых отчетов по МСФО. В конфигурационном файле input.txt
# указаны имя файла Excel и имя листа в нём, на котором содержатся данные,
# полученные из годовой отчетности компаний. Необходимыми для работы программы
# данными являются:
# - тикер;
# - прибыль;
# - номер эмитента на сайте www.e-disclosure.ru;
# - тип публикуемой отчетности на сайте www.e-disclosure.ru;
# - номер последней ссылки на странице финансовой отчетности эмитента.
#   Для запуска программы необходим Python 3.0 и библиотеки urllib3, certifi,
# xlrd, pandas.
#   При выполнении в командной строке:
#   python disclosure.py
#                    Copyright (c) 2019 - 2022 Логинов М.Д.
#  Разработчик: Логинов М.Д.
#  Модифицирован: 17 апреля 2022 г.
#
# *****************************************************************************

# -*- coding: utf-8 -*-
import os
import urllib3
import certifi
import pandas as pd
import openpyxl


# *****************************************************************************
#         Функция получения куки от сервера
# *****************************************************************************
def get_cookie():
    url = 'https://www.e-disclosure.ru/portal/files.aspx?id=9&type=4'
    http = urllib3.PoolManager(cert_reqs='CERT_REQUIRED',
                               ca_certs=certifi.where(),
                               timeout=urllib3.Timeout(connect=1.0, read=2.0))
    headers1 = {'Cookie': '', 'User-Agent':
                'Mozilla/5.0 (X11; Linux x86_64; rv:83.0) Gecko/20100101 ' +
                'Firefox/83.0'}
    resp = http.request('GET', url, headers=headers1, retries=False)
    return resp.getheaders()['Set-Cookie']


# *****************************************************************************
#         Функция получения данных из Интернета и их анализа
# *****************************************************************************
def read_new_data(id, typeId, cookies):
    url = 'https://www.e-disclosure.ru/portal/files.aspx?id=0&type='\
          + str(typeId)
    url1 = url.replace('0', str(id))
    http = urllib3.PoolManager(cert_reqs='CERT_REQUIRED',
                               ca_certs=certifi.where(),
                               timeout=urllib3.Timeout(connect=1.0, read=2.0))
    headers1 = {'Cookie': cookies, 'User-Agent':
                'Mozilla/5.0 (X11; Linux x86_64; rv:83.0) Gecko/20100101 ' +
                'Firefox/83.0'}
    resp = http.request('GET', url1, headers=headers1, retries=False)
    response = resp.data.decode('utf-8')
    f = str(response).split('\\r')
    # Следующая строка необходима для отладки offline
    # f=open('temp.html','r',encoding='utf-8')
    for line in f:
        if line.count('Fileid') > 0:
            str1 = line.split('Fileid=')
            str2 = str1[1].split('"')
            str3 = str2[0]
            break
    # f.close()
    try:
        return int(str3)
    except UnboundLocalError:
        return 0


# *****************************************************************************
#                  Функция сохранения отчета на диск
# *****************************************************************************
def save_report(id, cookies, directory):
    url = 'https://www.e-disclosure.ru/portal/FileLoad.ashx?Fileid=' + str(id)
    http = urllib3.PoolManager(cert_reqs='CERT_REQUIRED',
                               ca_certs=certifi.where(),
                               timeout=urllib3.Timeout(connect=1.0, read=2.0))
    headers1 = {'Cookie': cookies, 'User-Agent':
                'Mozilla/5.0 (X11; Linux x86_64; rv:83.0) Gecko/20100101 ' +
                'Firefox/83.0'}
    resp = http.request('GET', url, headers=headers1, retries=False)
    f = open('tmp', 'wb')
    f.write(resp.data)
    f.close()
    fileName = resp.headers['Content-Disposition'].split('=')[1].strip('"')
    os.rename('tmp', directory + fileName)
    return 0


# *****************************************************************************
#               Функция чтения исходных данных из файла Excel
# *****************************************************************************
def read_input(excelFile, sheetName):
    data = pd.read_excel(excelFile, sheetName, usecols=[1, 4, 15, 16, 17])
    data.set_axis(['ticker', 'profit', 'id', 'type', 'last'], axis=1,
                  inplace=True)
    data['profit'] = data['profit'].fillna(0)
    data = data[data['profit'] == 0]
    del data['profit']
    data['id'] = data['id'].fillna(0)
    data = data[data['id'] > 0]
    data['id'] = data['id'].astype('uint32')
    data['type'] = data['type'].fillna(4)
    data['type'] = data['type'].astype('uint8')
    data['last'] = data['last'].astype('uint32')
    data = data.reset_index(drop=True)
    return data


# *****************************************************************************
#            Функция чтения настроек из конфигурационного файла
#
# Результат: param[0] - путь к файлу Excel с результатами анализа отчетности по
#                       МСФО;
#            param[1] - имя листа с данными.
# *****************************************************************************
def read_configuration():
    f = open('input.txt', 'r', encoding='utf-8')
    param = []
    for line in f:
        param.append(line)
    f.close()
    return param[0][:-1], param[1][:-1]


# *****************************************************************************
# Функция изменения числового значения номера последней ссылки на странице
# финансовой отчетности указанного эмитента
#
# Результат: номер следующей строки на указанном листе в файле Excel
# *****************************************************************************
def change_cell(excelFile, sheetName, ticker, lastId, initialRow):
    workSheet = excelFile[sheetName]
    i = initialRow
    while workSheet.cell(row=i, column=2).value != ticker:
        i = i + 1
    workSheet.cell(row=i, column=18).value = lastId
    return i + 1


# *****************************************************************************
#                       Текст основной программы
# *****************************************************************************
excelFileName, sheetName = read_configuration()
data = read_input(excelFileName, sheetName)
excelFile = openpyxl.load_workbook(excelFileName)
n = 2
print('START')
siteCookie = get_cookie()
for i in range(len(data)):
    newData = read_new_data(data.loc[i, 'id'], data.loc[i, 'type'], siteCookie)
    if data.loc[i, 'last'] != newData and newData != 0:
        dir = data.loc[i, 'ticker'] + ';' + str(newData)
        print(dir)
        os.mkdir(dir)
        dir = os.getcwd() + '/' + dir + '/'
        save_report(newData, siteCookie, dir)
        n = change_cell(excelFile, sheetName, data.loc[i, 'ticker'], newData,
                        n)
print('STOP')
excelFile.save(excelFileName)